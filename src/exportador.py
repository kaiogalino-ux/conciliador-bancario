"""Geração do arquivo resultado/Resultado.xlsx.

Regra revisada em 2026-07-23 (ver docs/HISTORICO_DECISOES.md): o arquivo
passou a ter **duas abas**, "Resumo" e "Base Detalhada" — reproduzindo
fielmente o layout do arquivo-modelo `resultado/Modelo_principal_
conciliacao_status.xlsx` (analisado e usado como referência de estrutura,
cores, tipografia e ícones; nunca lido em tempo de execução nem
sobrescrito). Isso substitui a regra anterior de aba única "Resultado"
(2026-07-10-b).

- "Resumo": título, subtítulo com o período conciliado, o painel de 5 cards
  executivos (Total na Gestão, Total no Banco, Conciliado, Revisão Manual,
  Somente no Banco) e a tabela "Itens pendentes de análise".
- "Base Detalhada": título, cabeçalho com os nomes originais das colunas e
  todas as linhas do resultado, sem nenhuma coluna/linha removida, seguidas
  do rodapé (critério de lote + data/hora de geração).

Os 5 cards do painel são formas do Excel (retângulo arredondado + ícone),
não células — o openpyxl não permite criar esse tipo de forma. Por isso eles
são injetados depois, por `src/exportador_shapes.py`, a partir de um template
extraído uma única vez do arquivo-modelo (nunca com valores fixos: os 5
textos são sempre recalculados a partir do `resultado` desta execução). Esta
função (e todo o resto deste módulo) nunca decide nem altera nenhum valor,
Status, Tipo Conciliação, Motivo ou Observação — apenas lê o DataFrame já
concluído por `src/conciliador.py` e formata a apresentação. `conciliar()`
sempre roda antes e por completo; esta camada só é chamada depois.
"""

import logging
from datetime import date, datetime
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from src.conciliador import (
    COLUNAS_RESULTADO,
    STATUS_CONCILIADO,
    STATUS_REVISAO_MANUAL,
    STATUS_SOMENTE_BANCO,
)
from src.exportador_shapes import injetar_cards_e_icones
from src.ia_revisor import COLUNAS_IA

NOME_ABA_RESUMO = "Resumo"
NOME_ABA_BASE_DETALHADA = "Base Detalhada"

# Textos fixos do painel/seções — usados tanto na escrita quanto nos testes
# (ver tests/test_exportador_layout.py), para nunca divergir entre os dois.
# Os 5 títulos dos cards ficam bem aqui só como referência/documentação: o
# texto de fato exibido já está no template (src/assets/painel_visual/), pois
# os cards são formas do Excel, não células.
TITULO_PAINEL = "Quadro detalhado de conciliação"
TITULO_CARTAO_ERP = "TOTAL NA GESTÃO"
TITULO_CARTAO_BANCO = "TOTAL NO BANCO"
TITULO_CARTAO_CONCILIADO = "CONCILIADO"
TITULO_CARTAO_REVISAO_MANUAL = "REVISÃO MANUAL"
TITULO_CARTAO_SOMENTE_BANCO = "SOMENTE NO BANCO"

TITULO_PENDENCIAS = "Itens pendentes de análise"
MSG_SEM_PENDENCIAS = "Nenhum item pendente de análise."
MOTIVO_PADRAO_PENDENCIA = "Sem justificativa automática disponível"
COLUNAS_PENDENCIAS = [
    "Data",
    "Origem",
    "Favorecido ou descrição",
    "Valor na Gestão",
    "Valor no banco",
    "Status",
    "Motivo",
]

TITULO_BASE_DETALHADA = "Base detalhada completa"

TEXTO_RODAPE_CRITERIO = (
    "Critério: nos lotes consolidados, foram somados apenas os valores individuais de cada origem, "
    "sem repetição do total do lote."
)

# Formatos numéricos — moeda com seção negativa em vermelho, igual ao
# arquivo-modelo (sintaxe de format code do Excel: "," é separador de milhar
# e "." o decimal na própria sintaxe do código; o Excel exibe com os
# separadores do idioma do usuário, por isso já aparece como "R$ 1.234,56").
FORMATO_MOEDA = r'\R\$\ #,##0.00;[Red]\-\R\$\ #,##0.00'
FORMATO_DATA = "DD/MM/YYYY"
FORMATO_PERCENTUAL = "0.0%"
FORMATO_INTEIRO = "0"

COLUNAS_MOEDA = {"Valor ERP", "Valor Banco", "Possível Valor Banco"}
COLUNAS_DATA = {
    "Data ERP Usada", "Data Banco", "Data de Compensação Original",
    "Vencimento Original", "Possível Data Banco",
}
COLUNAS_PERCENTUAL = {"Confiança IA"}
COLUNAS_INTEIRO = {"Diferença de Dias"}

# Colunas que `conciliar()` continua produzindo normalmente (nada muda na
# lógica de conciliação), mas que deixaram de ser exportadas para a aba
# "Base Detalhada" por pedido explícito do usuário (2026-07-24 — ver
# docs/HISTORICO_DECISOES.md). Continuam disponíveis: no DataFrame em memória
# (para quem consumir `conciliar()` diretamente) e, quando aplicável, no
# arquivo de log do dia. "Motivo Revisão"/"Motivo Não Conciliado" continuam
# resumidos na coluna "Motivo" da tabela "Itens pendentes de análise" (aba
# "Resumo"), então essa informação não desaparece do arquivo, só sai da
# tabela detalhada.
COLUNAS_OCULTAS_BASE_DETALHADA = [
    "Data de Compensação Original",
    "Motivo Revisão",
    "Diferença de Dias",
    "ID Lote",
    "Possível Data Banco",
    "Possível Valor Banco",
    "Possível Descrição Banco",
    "Status do Possível Banco",
    "Decisão IA",
    "Confiança IA",
    "Motivo IA",
    "Validação IA",
    "Modelo IA",
]

# Paleta e tipografia extraídas do arquivo-modelo.
FONTE_PADRAO = "Arial"
COR_NAVY = "12376B"
COR_SUBTITULO = "6B6B6B"
COR_VERDE_FUNDO = "E2F0D9"
COR_VERDE_TEXTO = "375623"
COR_LARANJA_FUNDO = "FFF2CC"
COR_LARANJA_TEXTO = "7F6000"
COR_VERMELHO_FUNDO = "F4CCCC"
COR_VERMELHO_TEXTO = "9C0006"
COR_CABECALHO_TABELA = "DCE8F3"
COR_LINHA_ALTERNADA = "F4F7F9"
COR_BORDA_TABELA = "D9D9D9"

_BORDA_FINA = Border(
    left=Side(style="thin", color=COR_BORDA_TABELA),
    right=Side(style="thin", color=COR_BORDA_TABELA),
    top=Side(style="thin", color=COR_BORDA_TABELA),
    bottom=Side(style="thin", color=COR_BORDA_TABELA),
)

# --------------------------------------------------------------------------
# Layout do "Resumo" (linhas/colunas em EMU absolutos no template de cards —
# ver src/assets/painel_visual/ — por isso as alturas de linha 1-8 abaixo
# precisam bater com o arquivo-modelo para os cards ficarem alinhados).
# --------------------------------------------------------------------------
LINHA_TITULO = 1
LINHA_SUBTITULO = 2
LINHA_CARDS_INICIO = 4
LINHA_CARDS_FIM = 8
ALTURA_LINHA_TITULO = 23.1
ALTURA_LINHA_SUBTITULO = 17.1
ALTURA_LINHA_ESPACADOR_CARDS = 9.0
ALTURA_LINHA_CARDS = 15.0
ALTURA_LINHA_TITULO_PENDENCIAS = 21.0
ALTURA_LINHA_ESPACADOR_PENDENCIAS = 6.95
ALTURA_LINHA_CABECALHO_PENDENCIAS = 20.1
ALTURA_LINHA_DADOS_PENDENCIAS = 30.0

LARGURAS_COLUNAS_RESUMO = [12.7109375, 18.7109375, 42.7109375, 16.7109375, 13.0, 19.7109375, 48.7109375]

# Largura das colunas da "Base Detalhada", na mesma ordem de COLUNAS_RESULTADO
# + COLUNAS_IA (extraída do arquivo-modelo) — `None` mantém a largura padrão.
_LARGURAS_BASE_DETALHADA_MODELO = [
    15.7109375, 22.7109375, 20.7109375, 18.7109375, 15.7109375, 16.7109375, None,
    32.7109375, 38.7109375, 42.7109375, 19.7109375, 22.7109375, 34.7109375, None,
    36.7109375, 15.7109375, 16.7109375, 18.7109375, 19.7109375, 38.7109375, 22.7109375,
    18.7109375, None, 14.7109375, 38.7109375, 20.7109375, 18.7109375,
]
_ORDEM_COLUNAS_MODELO = COLUNAS_RESULTADO + COLUNAS_IA
LARGURA_BASE_DETALHADA_POR_COLUNA = dict(zip(_ORDEM_COLUNAS_MODELO, _LARGURAS_BASE_DETALHADA_MODELO))


def exportar_resultado(
    resultado: pd.DataFrame,
    caminho_saida: Path,
    logger: logging.Logger,
    periodo_inicial=None,
    periodo_final=None,
) -> None:
    """Grava `resultado` em `caminho_saida`, com as abas "Resumo" e "Base
    Detalhada" (ver docstring do módulo).

    `periodo_inicial`/`periodo_final` (opcionais) são o período já detectado
    por `src/leitor_erp.py` (ver main.py) — usados só para o subtítulo do
    painel; quando ausentes, o período é estimado a partir das datas do
    próprio `resultado`. Nenhum dos dois parâmetros participa da conciliação.
    """
    caminho_saida.parent.mkdir(parents=True, exist_ok=True)

    formatar_resultado_excel(
        caminho_saida, resultado, logger, periodo_inicial=periodo_inicial, periodo_final=periodo_final
    )

    logger.info(f"Resultado exportado para: {caminho_saida}")
    logger.info(f"Abas '{NOME_ABA_RESUMO}' e '{NOME_ABA_BASE_DETALHADA}': {len(resultado)} linha(s) na base.")


def formatar_resultado_excel(
    caminho: Path,
    resultado: pd.DataFrame,
    logger: logging.Logger,
    periodo_inicial=None,
    periodo_final=None,
) -> None:
    """Constrói e salva o workbook inteiro (Resumo + Base Detalhada).

    Chamada só depois que `conciliar()` já terminou — não decide nem altera
    nenhum dado de conciliação, só lê `resultado` e formata a apresentação.
    Os cards do painel (formas do Excel) são injetados depois de salvar, por
    `src/exportador_shapes.py`.
    """
    workbook = Workbook()
    ws_resumo = workbook.active
    ws_resumo.title = NOME_ABA_RESUMO
    ws_base = workbook.create_sheet(NOME_ABA_BASE_DETALHADA)

    metricas_financeiras = _calcular_metricas_financeiras(resultado)
    metricas_status = _calcular_metricas_status(resultado)
    pendentes = _montar_linhas_pendentes(resultado)
    p_inicial, p_final = _periodo_relatorio(resultado, periodo_inicial, periodo_final)
    agora = datetime.now()

    _formatar_resumo(ws_resumo, pendentes, p_inicial, p_final)
    _formatar_base_detalhada(ws_base, resultado, agora)

    workbook.save(caminho)

    valores_cards = {
        "total_erp": _formatar_moeda_texto(metricas_financeiras["total_erp"]),
        "total_banco": _formatar_moeda_texto(metricas_financeiras["total_banco"]),
        "conciliados": _texto_indicador(*metricas_status["conciliados"]),
        "revisao_manual": _texto_indicador(*metricas_status["revisao"]),
        "somente_banco": _texto_indicador(*metricas_status["somente_banco"]),
    }
    injetar_cards_e_icones(caminho, NOME_ABA_RESUMO, NOME_ABA_BASE_DETALHADA, valores_cards)

    logger.info(
        f"Painel: Total ERP {valores_cards['total_erp']}, Total banco {valores_cards['total_banco']}, "
        f"{len(pendentes)} item(ns) pendente(s) de análise."
    )


# --------------------------------------------------------------------------
# Sanitização de valores (nunca deixa "nan"/"None"/"NaT"/string vazia visível)
# --------------------------------------------------------------------------

def _valor_celula(valor):
    """Converte um valor do DataFrame para algo seguro de gravar numa célula.

    Tipos numpy viram tipos nativos do Python (openpyxl não reconhece
    confiavelmente `numpy.generic`); NaN/NaT/None/string vazia viram `None`
    (célula em branco) — nunca "nan"/"None"/"NaT" como texto visível.
    """
    if valor is None:
        return None

    if isinstance(valor, (datetime, date)):
        try:
            if pd.isna(valor):
                return None
        except (TypeError, ValueError):
            pass
        return valor

    if isinstance(valor, np.generic):
        valor = valor.item()

    if isinstance(valor, str):
        texto = valor.strip()
        if not texto or texto.lower() in {"nan", "none", "nat"}:
            return None
        return texto

    try:
        if pd.isna(valor):
            return None
    except (TypeError, ValueError):
        pass

    return valor


def _e_vazio(valor) -> bool:
    return _valor_celula(valor) is None


def _formatar_data_texto(data) -> str:
    if isinstance(data, (datetime, date)):
        return data.strftime("%d/%m/%Y")
    return str(data)


def _formatar_percentual_br(fracao: float) -> str:
    return f"{fracao * 100:.1f}".replace(".", ",") + "%"


def _formatar_moeda_texto(valor: float) -> str:
    """Formata um valor como texto "R$ 1.234,56" — usado só para o texto
    estático dos cards (formas do Excel, sem número/formatação nativa)."""
    texto = f"{valor:,.2f}"
    texto = texto.replace(",", "§").replace(".", ",").replace("§", ".")
    return f"R$ {texto}"


def _texto_indicador(qtd: int, pct: float) -> str:
    """Ex.: "151  (93,8%)" — mesmo formato do texto estático dos cards de
    status no arquivo-modelo."""
    return f"{qtd}  ({_formatar_percentual_br(pct)})"


def _para_date(valor):
    if isinstance(valor, datetime):
        return valor.date()
    return valor


# --------------------------------------------------------------------------
# Métricas do painel — calculadas em Python a partir do DataFrame final, sem
# nenhuma fórmula/tabela auxiliar no Excel.
# --------------------------------------------------------------------------

def _calcular_metricas_financeiras(resultado: pd.DataFrame) -> dict:
    """Totais consolidados do painel (cards "Total na Gestão" e "Total no Banco").

    Dentro de um lote NET EMP (ver `_resolver_lote_net_empr` em
    `src/conciliador.py`), cada linha do ERP também carrega, só para contexto
    visual, o total consolidado do banco do grupo em "Valor Banco" — e cada
    linha do banco carrega o total consolidado do ERP do grupo em "Valor
    ERP". Somar a coluna inteira sem filtro contaria esse total do grupo uma
    vez a mais por lançamento irmão. Por isso, o total do ERP soma apenas as
    linhas cuja "Origem" é de fato um lançamento do ERP (`ERP`/`ERP+Banco`), e
    o total do banco soma apenas as linhas cuja "Origem" é de fato um
    lançamento do banco (`Banco`/`ERP+Banco`) — usando exatamente o marcador
    ("Origem") que o conciliador já produz, sem nenhuma regra nova de lote.
    """
    total_erp = 0.0
    total_banco = 0.0
    if "Valor ERP" in resultado.columns and "Origem" in resultado.columns:
        mascara_erp = resultado["Origem"].isin(("ERP", "ERP+Banco"))
        total_erp = round(float(resultado.loc[mascara_erp, "Valor ERP"].dropna().sum()), 2)
    elif "Valor ERP" in resultado.columns:
        total_erp = round(float(resultado["Valor ERP"].dropna().sum()), 2)

    if "Valor Banco" in resultado.columns and "Origem" in resultado.columns:
        mascara_banco = resultado["Origem"].isin(("Banco", "ERP+Banco"))
        total_banco = round(float(resultado.loc[mascara_banco, "Valor Banco"].dropna().abs().sum()), 2)
    elif "Valor Banco" in resultado.columns:
        total_banco = round(float(resultado["Valor Banco"].dropna().abs().sum()), 2)

    return {"total_erp": total_erp, "total_banco": total_banco}


def _calcular_metricas_status(resultado: pd.DataFrame) -> dict:
    """Conciliados / Revisão manual / Somente banco (cards do painel) — mais
    "não conciliados" (Somente banco + Não encontrado no banco + qualquer
    outro status), mantido internamente para a soma de percentuais fechar em
    100%. Nunca altera o Status original: só lê e conta.
    """
    total = len(resultado)
    if total == 0 or "Status" not in resultado.columns:
        vazio = (0, 0.0)
        return {"conciliados": vazio, "revisao": vazio, "somente_banco": vazio, "nao_conciliados": vazio, "total": total}

    status = resultado["Status"]
    qtd_conciliados = int((status == STATUS_CONCILIADO).sum())
    qtd_revisao = int((status == STATUS_REVISAO_MANUAL).sum())
    qtd_somente_banco = int((status == STATUS_SOMENTE_BANCO).sum())
    qtd_nao_conciliados = total - qtd_conciliados - qtd_revisao

    return {
        "conciliados": (qtd_conciliados, qtd_conciliados / total),
        "revisao": (qtd_revisao, qtd_revisao / total),
        "somente_banco": (qtd_somente_banco, qtd_somente_banco / total),
        "nao_conciliados": (qtd_nao_conciliados, qtd_nao_conciliados / total),
        "total": total,
    }


def _periodo_relatorio(resultado: pd.DataFrame, periodo_inicial=None, periodo_final=None):
    """Usa o período já detectado por `ler_erp()` (passado por main.py) quando
    disponível; senão estima a partir do menor/maior data presentes no próprio
    resultado (nunca inventa um período)."""
    if periodo_inicial is not None and periodo_final is not None:
        return periodo_inicial, periodo_final

    datas = []
    for coluna in ("Data ERP Usada", "Data Banco"):
        if coluna in resultado.columns:
            datas.extend(_para_date(valor) for valor in resultado[coluna] if not _e_vazio(valor))

    if not datas:
        return None, None
    return min(datas), max(datas)


# --------------------------------------------------------------------------
# Itens pendentes de análise
# --------------------------------------------------------------------------

def _origem_visivel(origem) -> str:
    return {"ERP": "Gestão", "Banco": "Banco", "ERP+Banco": "Gestão / Banco"}.get(origem, origem or "")


def _data_pendencia(linha: dict):
    origem = linha.get("Origem")
    data_erp = linha.get("Data ERP Usada")
    data_banco = linha.get("Data Banco")
    if origem in ("ERP", "ERP+Banco") and not _e_vazio(data_erp):
        return data_erp
    if not _e_vazio(data_banco):
        return data_banco
    return data_erp


def _favorecido_pendencia(linha: dict) -> str:
    for campo in ("Favorecido", "Descrição ERP", "Descrição Banco"):
        texto = _valor_celula(linha.get(campo))
        if texto:
            return texto
    for campo in ("Motivo Revisão", "Motivo Não Conciliado", "Observações"):
        texto = _valor_celula(linha.get(campo))
        if texto:
            return texto
    return ""


def _motivo_pendencia(linha: dict) -> str:
    for campo in ("Motivo Revisão", "Motivo Não Conciliado", "Observações"):
        texto = _valor_celula(linha.get(campo))
        if texto:
            return texto
    return MOTIVO_PADRAO_PENDENCIA


def _montar_linhas_pendentes(resultado: pd.DataFrame) -> list:
    """Só os registros que realmente precisam de atenção humana (tudo que não
    é `Conciliado`) — nunca inclui um registro já conciliado, e nunca altera o
    Status/Motivo/Observações originais, só os lê para montar a linha visual."""
    if "Status" not in resultado.columns or len(resultado) == 0:
        return []

    mascara = resultado["Status"] != STATUS_CONCILIADO
    pendentes = []
    for linha in resultado.loc[mascara].to_dict("records"):
        valor_erp = linha.get("Valor ERP")
        valor_banco = linha.get("Valor Banco")
        pendentes.append({
            "Data": _data_pendencia(linha),
            "Origem": _origem_visivel(linha.get("Origem")),
            "Favorecido ou descrição": _favorecido_pendencia(linha),
            "Valor na Gestão": None if _e_vazio(valor_erp) else valor_erp,
            "Valor no banco": None if _e_vazio(valor_banco) else abs(valor_banco),
            "Status": linha.get("Status"),
            "Motivo": _motivo_pendencia(linha),
        })
    return pendentes


# --------------------------------------------------------------------------
# Escrita das seções (célula a célula — os cards são injetados depois, como
# formas, por src/exportador_shapes.py)
# --------------------------------------------------------------------------

def _cel(ws, linha, coluna, valor=None, fonte=None, preenchimento=None, alinhamento=None, borda=None, formato=None):
    celula = ws.cell(row=linha, column=coluna, value=valor)
    if fonte is not None:
        celula.font = fonte
    if preenchimento is not None:
        celula.fill = preenchimento
    if alinhamento is not None:
        celula.alignment = alinhamento
    if borda is not None:
        celula.border = borda
    if formato is not None:
        celula.number_format = formato
    return celula


def _mesclar(ws, linha_ini, col_ini, linha_fim, col_fim):
    if col_fim < col_ini:
        col_fim = col_ini
    ws.merge_cells(start_row=linha_ini, start_column=col_ini, end_row=linha_fim, end_column=col_fim)


def _formatar_resumo(ws, pendentes: list, periodo_inicial, periodo_final) -> None:
    ws.sheet_view.showGridLines = False
    ws.sheet_view.zoomScale = 90

    fonte_titulo = Font(name=FONTE_PADRAO, size=16, bold=True, color=COR_NAVY)
    fonte_subtitulo = Font(name=FONTE_PADRAO, size=9, color=COR_SUBTITULO)

    ws.row_dimensions[LINHA_TITULO].height = ALTURA_LINHA_TITULO
    _mesclar(ws, LINHA_TITULO, 1, LINHA_TITULO, 24)
    _cel(ws, LINHA_TITULO, 1, TITULO_PAINEL, fonte=fonte_titulo, alinhamento=Alignment(vertical="center"))

    ws.row_dimensions[LINHA_SUBTITULO].height = ALTURA_LINHA_SUBTITULO
    if periodo_inicial is not None and periodo_final is not None:
        subtitulo = (
            f"Conferência entre a Gestão (ERP) e o banco  •  "
            f"{_formatar_data_texto(periodo_inicial)} a {_formatar_data_texto(periodo_final)}"
        )
    else:
        subtitulo = "Conferência entre a Gestão (ERP) e o banco"
    _mesclar(ws, LINHA_SUBTITULO, 1, LINHA_SUBTITULO, 24)
    _cel(ws, LINHA_SUBTITULO, 1, subtitulo, fonte=fonte_subtitulo, alinhamento=Alignment(vertical="center"))

    ws.row_dimensions[3].height = ALTURA_LINHA_ESPACADOR_CARDS
    for linha_card in range(LINHA_CARDS_INICIO, LINHA_CARDS_FIM + 1):
        ws.row_dimensions[linha_card].height = ALTURA_LINHA_CARDS

    linha = LINHA_CARDS_FIM + 1
    linha = _escrever_pendencias(ws, linha, pendentes)

    for indice, largura in enumerate(LARGURAS_COLUNAS_RESUMO, start=1):
        ws.column_dimensions[get_column_letter(indice)].width = largura

    ws.page_setup.orientation = "landscape"
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins.left = 0.4
    ws.page_margins.right = 0.4
    ws.page_margins.top = 0.6
    ws.page_margins.bottom = 0.6
    ws.print_area = f"A1:G{max(linha - 1, 21)}"


def _escrever_pendencias(ws, linha, pendentes: list) -> int:
    fonte_titulo_secao = Font(name=FONTE_PADRAO, size=12, bold=True, color=COR_NAVY)

    ws.row_dimensions[linha].height = ALTURA_LINHA_TITULO_PENDENCIAS
    _cel(ws, linha, 1, TITULO_PENDENCIAS, fonte=fonte_titulo_secao,
         alinhamento=Alignment(horizontal="left", indent=3))
    linha += 1

    ws.row_dimensions[linha].height = ALTURA_LINHA_ESPACADOR_PENDENCIAS
    linha += 1

    if not pendentes:
        _mesclar(ws, linha, 1, linha, len(COLUNAS_PENDENCIAS))
        _cel(
            ws, linha, 1, MSG_SEM_PENDENCIAS,
            fonte=Font(name=FONTE_PADRAO, size=10, bold=True, color=COR_VERDE_TEXTO),
            preenchimento=PatternFill("solid", fgColor=COR_VERDE_FUNDO),
            alinhamento=Alignment(horizontal="center", vertical="center"),
        )
        return linha + 1

    fonte_cabecalho = Font(name=FONTE_PADRAO, size=8, bold=True, color=COR_NAVY)
    preenchimento_cabecalho = PatternFill("solid", fgColor=COR_CABECALHO_TABELA)
    alinhamento_cabecalho = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.row_dimensions[linha].height = ALTURA_LINHA_CABECALHO_PENDENCIAS
    for indice, nome_coluna in enumerate(COLUNAS_PENDENCIAS, start=1):
        _cel(ws, linha, indice, nome_coluna, fonte=fonte_cabecalho, preenchimento=preenchimento_cabecalho,
             alinhamento=alinhamento_cabecalho, borda=_BORDA_FINA)
    linha += 1

    cores_status = {
        STATUS_REVISAO_MANUAL: (COR_LARANJA_FUNDO, COR_LARANJA_TEXTO),
        STATUS_CONCILIADO: (COR_VERDE_FUNDO, COR_VERDE_TEXTO),
    }

    for offset, item in enumerate(pendentes):
        linha_atual = linha + offset
        ws.row_dimensions[linha_atual].height = ALTURA_LINHA_DADOS_PENDENCIAS
        cor_fundo, cor_texto = cores_status.get(item["Status"], (COR_VERMELHO_FUNDO, COR_VERMELHO_TEXTO))
        for indice, nome_coluna in enumerate(COLUNAS_PENDENCIAS, start=1):
            valor = _valor_celula(item.get(nome_coluna))
            fonte = Font(name=FONTE_PADRAO, size=8)
            alinhamento = Alignment(horizontal="left", vertical="center", wrap_text=True)
            formato = None
            preenchimento = None

            if nome_coluna in ("Valor na Gestão", "Valor no banco"):
                formato = FORMATO_MOEDA
                alinhamento = Alignment(horizontal="right", vertical="center", wrap_text=True)
            elif nome_coluna == "Data":
                formato = FORMATO_DATA
                alinhamento = Alignment(horizontal="center", vertical="center", wrap_text=True)
            elif nome_coluna in ("Origem", "Status"):
                alinhamento = Alignment(horizontal="center", vertical="center", wrap_text=True)

            if nome_coluna == "Status":
                preenchimento = PatternFill("solid", fgColor=cor_fundo)
                fonte = Font(name=FONTE_PADRAO, size=8, bold=True, color=cor_texto)

            _cel(ws, linha_atual, indice, valor, fonte=fonte, preenchimento=preenchimento,
                 alinhamento=alinhamento, borda=_BORDA_FINA, formato=formato)

    return linha + len(pendentes)


def _formato_para_coluna(nome_coluna):
    if nome_coluna in COLUNAS_MOEDA:
        return FORMATO_MOEDA
    if nome_coluna in COLUNAS_DATA:
        return FORMATO_DATA
    if nome_coluna in COLUNAS_PERCENTUAL:
        return FORMATO_PERCENTUAL
    if nome_coluna in COLUNAS_INTEIRO:
        return FORMATO_INTEIRO
    return None


def _formatar_base_detalhada(ws, resultado: pd.DataFrame, agora) -> None:
    """Escreve, na aba "Base Detalhada": título (linha 1), cabeçalho com os
    nomes das colunas visíveis (linha 2) e as linhas de `resultado` a partir
    da linha 3 — nunca reordena nem altera valor/linha, só omite as colunas
    de `COLUNAS_OCULTAS_BASE_DETALHADA` (pedido explícito do usuário,
    2026-07-24) — seguidas do rodapé."""
    ws.sheet_view.showGridLines = False
    ws.sheet_view.zoomScale = 80

    fonte_titulo_secao = Font(name=FONTE_PADRAO, size=12, bold=True, color=COR_NAVY)
    ws.row_dimensions[1].height = ALTURA_LINHA_TITULO
    _cel(ws, 1, 1, TITULO_BASE_DETALHADA, fonte=fonte_titulo_secao, alinhamento=Alignment(horizontal="left", indent=3))

    colunas = [nome for nome in resultado.columns if nome not in COLUNAS_OCULTAS_BASE_DETALHADA]
    resultado_visivel = resultado[colunas]
    total_colunas = len(colunas)
    linha_cabecalho = 2

    fonte_cabecalho = Font(name=FONTE_PADRAO, size=8, bold=True, color=COR_NAVY)
    preenchimento_cabecalho = PatternFill("solid", fgColor=COR_CABECALHO_TABELA)
    alinhamento_cabecalho = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.row_dimensions[linha_cabecalho].height = 38.1
    for indice_coluna, nome_coluna in enumerate(colunas, start=1):
        _cel(ws, linha_cabecalho, indice_coluna, nome_coluna, fonte=fonte_cabecalho,
             preenchimento=preenchimento_cabecalho, alinhamento=alinhamento_cabecalho, borda=_BORDA_FINA)

    alinhamento_texto = Alignment(horizontal="left", vertical="center", wrap_text=True)
    alinhamento_numerico = Alignment(horizontal="right", vertical="center")
    alinhamento_data = Alignment(horizontal="center", vertical="center")
    preenchimento_alternado = PatternFill("solid", fgColor=COR_LINHA_ALTERNADA)
    fonte_dado = Font(name=FONTE_PADRAO, size=8, color="000000")

    primeira_linha_dados = linha_cabecalho + 1
    total_linhas = len(resultado_visivel)

    for indice_linha, valores in enumerate(resultado_visivel.itertuples(index=False, name=None)):
        linha_atual = primeira_linha_dados + indice_linha
        preenchimento = preenchimento_alternado if indice_linha % 2 == 1 else None
        for indice_coluna, (nome_coluna, valor_bruto) in enumerate(zip(colunas, valores), start=1):
            valor = _valor_celula(valor_bruto)
            formato = _formato_para_coluna(nome_coluna)
            if formato == FORMATO_DATA:
                alinhamento = alinhamento_data
            elif formato in (FORMATO_MOEDA, FORMATO_PERCENTUAL, FORMATO_INTEIRO):
                alinhamento = alinhamento_numerico
            else:
                alinhamento = alinhamento_texto
            _cel(ws, linha_atual, indice_coluna, valor, fonte=fonte_dado, preenchimento=preenchimento,
                 alinhamento=alinhamento, borda=_BORDA_FINA, formato=formato)

    ultima_linha_dados = primeira_linha_dados + total_linhas - 1 if total_linhas else linha_cabecalho

    linha_rodape = ultima_linha_dados + 2
    _escrever_rodape(ws, linha_rodape, total_colunas, agora)

    _ajustar_largura_colunas_base(ws, colunas)
    _configurar_pagina_base(ws, linha_cabecalho, ultima_linha_dados, total_colunas, linha_rodape)


def _escrever_rodape(ws, linha, total_colunas, agora) -> None:
    fonte = Font(name="Calibri", size=9, italic=True, color=COR_SUBTITULO)
    texto_direita = f"Gerado em {agora.strftime('%d/%m/%Y às %H:%M')}"

    col_direita_ini = max(1, total_colunas - 2)
    col_esquerda_fim = max(1, col_direita_ini - 1)

    _mesclar(ws, linha, 1, linha, col_esquerda_fim)
    _cel(ws, linha, 1, TEXTO_RODAPE_CRITERIO, fonte=fonte,
         alinhamento=Alignment(horizontal="left", vertical="center", wrap_text=True))

    _mesclar(ws, linha, col_direita_ini, linha, max(total_colunas, col_direita_ini))
    _cel(ws, linha, col_direita_ini, texto_direita, fonte=fonte,
         alinhamento=Alignment(horizontal="right", vertical="center"))


def _ajustar_largura_colunas_base(ws, colunas: list) -> None:
    """Usa as larguras extraídas do arquivo-modelo (por nome de coluna, não
    por posição — nunca presume que todas as colunas do modelo estarão
    presentes). Para uma coluna fora dessa lista (ex.: DataFrame sintético de
    teste), calcula a largura a partir do próprio conteúdo, como antes."""
    for indice, nome_coluna in enumerate(colunas, start=1):
        letra = get_column_letter(indice)
        if nome_coluna in LARGURA_BASE_DETALHADA_POR_COLUNA:
            largura = LARGURA_BASE_DETALHADA_POR_COLUNA[nome_coluna]
            if largura is not None:
                ws.column_dimensions[letra].width = largura
            continue
        ws.column_dimensions[letra].width = min(max(len(str(nome_coluna)) + 2, 12), 60)


def _configurar_pagina_base(ws, linha_cabecalho, ultima_linha_dados, total_colunas, linha_rodape) -> None:
    # O cabeçalho ocupa as linhas 1 (título) e 2 (nomes das colunas); nada
    # abaixo da linha 3 pode ficar congelado (isso incluiria dados).
    ws.freeze_panes = "A3"

    letra_final = get_column_letter(max(total_colunas, 1))
    if ultima_linha_dados >= linha_cabecalho:
        ws.auto_filter.ref = f"A{linha_cabecalho}:{letra_final}{ultima_linha_dados}"

    ws.print_title_rows = "1:2"

    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins.left = 0.511811024
    ws.page_margins.right = 0.511811024
    ws.page_margins.top = 0.787401575
    ws.page_margins.bottom = 0.787401575
    ws.print_area = f"A1:{letra_final}{linha_rodape}"


def localizar_linha_com_texto(planilha, texto, coluna=1, limite=5000):
    """Procura, na coluna dada, a primeira linha cujo valor é exatamente
    `texto`. Usado pelos testes para nunca depender de números de linha fixos
    (a posição das seções muda conforme a quantidade de pendências)."""
    limite = min(limite, planilha.max_row)
    for linha in range(1, limite + 1):
        if planilha.cell(row=linha, column=coluna).value == texto:
            return linha
    return None
