"""Testes da camada de apresentação do Resultado.xlsx (src/exportador.py e
src/exportador_shapes.py).

Cobre só a formatação/layout (painel executivo com 5 cards, itens pendentes
de análise e base detalhada completa, agora em 2 abas — "Resumo" e "Base
Detalhada", reproduzindo o arquivo-modelo `resultado/Modelo_principal_
conciliacao_status.xlsx`) — nunca a lógica de conciliação, que já é protegida
pelos demais arquivos de teste. Todo cenário aqui usa `conciliar()` de
verdade (nunca fabrica um DataFrame de resultado à mão para os testes de
ponta a ponta), para garantir que a camada de apresentação nunca decide nem
altera Status/Tipo Conciliação/Motivo/Observações.

Os 5 cards do painel são formas do Excel (não células — ver
src/exportador_shapes.py), então os testes que os verificam leem o .xlsx
como zip e inspecionam o XML de `xl/drawings/drawing1.xml` diretamente.
"""

import zipfile
from datetime import date, datetime

import openpyxl
import pandas as pd
import pytest

from src.conciliador import (
    STATUS_CONCILIADO,
    STATUS_REVISAO_MANUAL,
    STATUS_SOMENTE_BANCO,
    conciliar,
)
from src.exportador import (
    COLUNAS_OCULTAS_BASE_DETALHADA,
    COLUNAS_PENDENCIAS,
    COR_LARANJA_FUNDO,
    COR_VERMELHO_FUNDO,
    MOTIVO_PADRAO_PENDENCIA,
    MSG_SEM_PENDENCIAS,
    NOME_ABA_BASE_DETALHADA,
    NOME_ABA_RESUMO,
    TITULO_BASE_DETALHADA,
    TITULO_CARTAO_BANCO,
    TITULO_CARTAO_CONCILIADO,
    TITULO_CARTAO_ERP,
    TITULO_CARTAO_REVISAO_MANUAL,
    TITULO_CARTAO_SOMENTE_BANCO,
    TITULO_PAINEL,
    TITULO_PENDENCIAS,
    _calcular_metricas_financeiras,
    _calcular_metricas_status,
    _formatar_moeda_texto,
    _montar_linhas_pendentes,
    _texto_indicador,
    exportar_resultado,
    localizar_linha_com_texto,
)
from src.exportador_shapes import PLACEHOLDERS
from tests.conftest import construir_df_banco, construir_df_erp


def _construir_cenario_misto():
    """Um pouco de cada status (Conciliado individual, Conciliado por lote NET
    EMP, Revisão Manual, Não encontrado no banco, Somente banco), para exercer
    o painel/pendências/base com todas as categorias visuais ao mesmo tempo."""
    df_erp = construir_df_erp([
        # Conciliado individual (par único).
        {"data_usada": date(2026, 5, 10), "valor": 353.00, "favorecido": "IB Extintores - NF 2037"},
        # Não encontrado no banco (nenhum lançamento bancário correspondente).
        {"data_usada": date(2026, 5, 13), "valor": 900.00, "favorecido": "Fornecedor XPTO Ltda"},
        # Revisão Manual: duplicidade de valor/data sem descrição suficiente.
        {"data_usada": date(2026, 5, 20), "valor": 250.00, "favorecido": "Serviço Diversos A"},
        {"data_usada": date(2026, 5, 20), "valor": 250.00, "favorecido": "Serviço Diversos A"},
        # Lote NET EMP (Salário/Folha) fechado por total direto.
        {"data_usada": date(2026, 5, 6), "valor": 1000.00, "favorecido": "Salario Fulano de Tal"},
        {"data_usada": date(2026, 5, 6), "valor": 1500.00, "favorecido": "Salario Beltrano da Silva"},
    ])
    df_banco = construir_df_banco([
        {"data": date(2026, 5, 10), "valor": -353.00, "favorecido": "PIX ENVIADO DES: IB EXTINTORES"},
        {"data": date(2026, 5, 20), "valor": -250.00, "favorecido": "PAGTO ELETRON COBRANCA XPTO"},
        {"data": date(2026, 5, 20), "valor": -250.00, "favorecido": "PAGTO ELETRON COBRANCA XPTO"},
        # Somente banco (nenhum ERP corresponde).
        {"data": date(2026, 5, 14), "valor": -150.00, "favorecido": "TARIFA BANCARIA XYZ"},
        # Lote NET EMP: banco genérico, total direto bate com os 2 ERP acima.
        {"data": date(2026, 5, 6), "valor": -2500.00, "favorecido": "PGTO SALARIO VIA NET EMP"},
    ])
    return df_erp, df_banco


@pytest.fixture
def resultado_misto(logger_silencioso):
    df_erp, df_banco = _construir_cenario_misto()
    return conciliar(df_erp, df_banco, logger_silencioso)


@pytest.fixture
def arquivo_misto(tmp_path, resultado_misto, logger_silencioso):
    caminho = tmp_path / "Resultado.xlsx"
    exportar_resultado(resultado_misto, caminho, logger_silencioso)
    return caminho, resultado_misto


@pytest.fixture
def planilha_mista(arquivo_misto):
    caminho, resultado = arquivo_misto
    workbook = openpyxl.load_workbook(caminho)
    return workbook, resultado, caminho


def _colunas_visiveis(resultado: pd.DataFrame) -> list:
    """Colunas que devem aparecer na aba "Base Detalhada" — todas as de
    `resultado` exceto as explicitamente ocultadas (pedido do usuário,
    2026-07-24, ver COLUNAS_OCULTAS_BASE_DETALHADA)."""
    return [nome for nome in resultado.columns if nome not in COLUNAS_OCULTAS_BASE_DETALHADA]


def _ler_drawing_cards(caminho) -> str:
    with zipfile.ZipFile(caminho) as z:
        return z.read("xl/drawings/drawing1.xml").decode("utf-8")


# --------------------------------------------------------------------------
# Estrutura geral: 2 abas (Resumo + Base Detalhada), arquivo abre sem erro.
# --------------------------------------------------------------------------

def test_gera_exatamente_as_2_abas_do_modelo(planilha_mista):
    workbook, _, _ = planilha_mista
    assert workbook.sheetnames == [NOME_ABA_RESUMO, NOME_ABA_BASE_DETALHADA] == ["Resumo", "Base Detalhada"]


def test_arquivo_zip_valido_e_abre_sem_erro(arquivo_misto):
    caminho, _ = arquivo_misto
    with zipfile.ZipFile(caminho) as z:
        assert z.testzip() is None

    workbook = openpyxl.load_workbook(caminho, read_only=True)
    assert workbook.sheetnames == ["Resumo", "Base Detalhada"]


def test_titulo_do_painel_presente_no_resumo(planilha_mista):
    workbook, _, _ = planilha_mista
    ws = workbook[NOME_ABA_RESUMO]
    assert localizar_linha_com_texto(ws, TITULO_PAINEL) == 1


# --------------------------------------------------------------------------
# Os 5 cards (formas do Excel — inspecionados no XML de drawing1.xml).
# --------------------------------------------------------------------------

def test_existem_as_5_formas_de_card_do_modelo(arquivo_misto):
    caminho, _ = arquivo_misto
    drawing = _ler_drawing_cards(caminho)
    for nome_forma in ("Card_Gestao", "Card_Banco", "Card_Conciliado", "Card_Revisao", "Card_SomenteBanco"):
        assert f'name="{nome_forma}"' in drawing, f"forma '{nome_forma}' não encontrada em drawing1.xml"


def test_titulos_dos_5_cards_presentes_no_drawing(arquivo_misto):
    caminho, _ = arquivo_misto
    drawing = _ler_drawing_cards(caminho)
    for titulo in (
        TITULO_CARTAO_ERP, TITULO_CARTAO_BANCO, TITULO_CARTAO_CONCILIADO,
        TITULO_CARTAO_REVISAO_MANUAL, TITULO_CARTAO_SOMENTE_BANCO,
    ):
        assert titulo in drawing


def test_nenhum_placeholder_do_template_sobra_no_drawing_final(arquivo_misto):
    caminho, _ = arquivo_misto
    drawing = _ler_drawing_cards(caminho)
    for placeholder in PLACEHOLDERS.values():
        assert placeholder not in drawing, f"placeholder {placeholder!r} não foi substituído"


def test_valores_dos_cards_batem_com_metricas_calculadas(arquivo_misto, resultado_misto):
    caminho, _ = arquivo_misto
    drawing = _ler_drawing_cards(caminho)

    metricas_fin = _calcular_metricas_financeiras(resultado_misto)
    metricas_status = _calcular_metricas_status(resultado_misto)

    assert _formatar_moeda_texto(metricas_fin["total_erp"]) in drawing
    assert _formatar_moeda_texto(metricas_fin["total_banco"]) in drawing
    assert _texto_indicador(*metricas_status["conciliados"]) in drawing
    assert _texto_indicador(*metricas_status["revisao"]) in drawing
    assert _texto_indicador(*metricas_status["somente_banco"]) in drawing


def test_valores_dos_cards_nao_sao_hardcoded_variam_com_o_resultado(tmp_path, logger_silencioso):
    """Gera 2 cenários com totais diferentes e confirma que os textos dos
    cards mudam de acordo — prova de que não há valor fixo extraído da
    planilha de exemplo."""
    df_erp_a = construir_df_erp([{"data_usada": date(2026, 5, 10), "valor": 111.00, "favorecido": "Fornecedor A"}])
    df_banco_a = construir_df_banco([{"data": date(2026, 5, 10), "valor": -111.00, "favorecido": "PIX FORNECEDOR A"}])
    resultado_a = conciliar(df_erp_a, df_banco_a, logger_silencioso)
    caminho_a = tmp_path / "a.xlsx"
    exportar_resultado(resultado_a, caminho_a, logger_silencioso)

    df_erp_b = construir_df_erp([{"data_usada": date(2026, 5, 10), "valor": 999.00, "favorecido": "Fornecedor B"}])
    df_banco_b = construir_df_banco([{"data": date(2026, 5, 10), "valor": -999.00, "favorecido": "PIX FORNECEDOR B"}])
    resultado_b = conciliar(df_erp_b, df_banco_b, logger_silencioso)
    caminho_b = tmp_path / "b.xlsx"
    exportar_resultado(resultado_b, caminho_b, logger_silencioso)

    drawing_a = _ler_drawing_cards(caminho_a)
    drawing_b = _ler_drawing_cards(caminho_b)

    assert _formatar_moeda_texto(111.00) in drawing_a
    assert _formatar_moeda_texto(999.00) in drawing_b
    assert _formatar_moeda_texto(111.00) not in drawing_b
    assert _formatar_moeda_texto(999.00) not in drawing_a


def test_icones_do_modelo_sao_reaproveitados(arquivo_misto):
    """Os 4 pares de ícone (PNG+SVG) do arquivo-modelo devem estar embutidos
    no .xlsx gerado — não uma aproximação (emoji/imagem genérica)."""
    caminho, _ = arquivo_misto
    with zipfile.ZipFile(caminho) as z:
        nomes = set(z.namelist())
    for indice in range(1, 9):
        extensao = "png" if indice % 2 == 1 else "svg"
        assert f"xl/media/image{indice}.{extensao}" in nomes


# --------------------------------------------------------------------------
# Cálculo das métricas (quantidades e percentuais).
# --------------------------------------------------------------------------

def test_metricas_financeiras_nao_duplicam_total_do_lote_net_emp(resultado_misto):
    metricas = _calcular_metricas_financeiras(resultado_misto)
    total_erp_esperado = 353.00 + 900.00 + 250.00 + 250.00 + 1000.00 + 1500.00
    total_banco_esperado = 353.00 + 250.00 + 250.00 + 150.00 + 2500.00
    assert metricas["total_erp"] == pytest.approx(total_erp_esperado, abs=0.01)
    assert metricas["total_banco"] == pytest.approx(total_banco_esperado, abs=0.01)


def test_metricas_status_incluem_somente_banco_e_somam_100_por_cento(resultado_misto):
    metricas = _calcular_metricas_status(resultado_misto)
    soma_qtd = metricas["conciliados"][0] + metricas["revisao"][0] + metricas["nao_conciliados"][0]
    assert soma_qtd == metricas["total"] == len(resultado_misto)

    soma_pct = metricas["conciliados"][1] + metricas["revisao"][1] + metricas["nao_conciliados"][1]
    assert soma_pct == pytest.approx(1.0, abs=1e-9)

    # "Somente banco" é um subconjunto de "não conciliados" (aqui, 1 dos 2).
    assert metricas["somente_banco"][0] == 1
    assert metricas["somente_banco"][0] <= metricas["nao_conciliados"][0]


def test_metricas_status_com_dataframe_vazio_nao_quebra():
    resultado = pd.DataFrame(columns=["Status", "Valor ERP", "Valor Banco", "Origem"])
    metricas = _calcular_metricas_status(resultado)
    assert metricas["total"] == 0
    for chave in ("conciliados", "revisao", "somente_banco", "nao_conciliados"):
        assert metricas[chave] == (0, 0.0)


# --------------------------------------------------------------------------
# Itens pendentes de análise: conteúdo e cores por status.
# --------------------------------------------------------------------------

def test_secao_pendencias_lista_so_o_que_nao_esta_conciliado(resultado_misto):
    pendentes = _montar_linhas_pendentes(resultado_misto)
    qtd_nao_conciliado_original = int((resultado_misto["Status"] != STATUS_CONCILIADO).sum())
    assert len(pendentes) == qtd_nao_conciliado_original
    for item in pendentes:
        assert item["Status"] != STATUS_CONCILIADO


def test_pendencias_tem_as_7_colunas_do_pedido(resultado_misto):
    pendentes = _montar_linhas_pendentes(resultado_misto)
    assert pendentes
    for item in pendentes:
        assert set(COLUNAS_PENDENCIAS) == set(item.keys())


def test_pendencias_nunca_mostra_motivo_vazio(resultado_misto):
    pendentes = _montar_linhas_pendentes(resultado_misto)
    for item in pendentes:
        assert item["Motivo"]
        assert item["Motivo"].strip().lower() not in {"nan", "none", "nat", ""}


def test_pendencias_sem_motivo_preenchido_usa_texto_padrao():
    resultado = pd.DataFrame([{
        "Status": STATUS_REVISAO_MANUAL, "Origem": "ERP", "Valor ERP": 10.0, "Valor Banco": None,
        "Motivo Revisão": None, "Motivo Não Conciliado": None, "Observações": None,
        "Favorecido": None, "Descrição ERP": None, "Descrição Banco": None,
        "Data ERP Usada": date(2026, 5, 1), "Data Banco": None,
    }])
    pendentes = _montar_linhas_pendentes(resultado)
    assert pendentes[0]["Motivo"] == MOTIVO_PADRAO_PENDENCIA


def test_cores_do_status_na_tabela_de_pendencias(planilha_mista):
    """Revisão Manual em amarelo, Somente banco em vermelho — sem alterar o
    texto original do Status."""
    workbook, resultado, _ = planilha_mista
    ws = workbook[NOME_ABA_RESUMO]

    linha_titulo = localizar_linha_com_texto(ws, TITULO_PENDENCIAS)
    linha_cabecalho = linha_titulo + 2  # título, espaçador, cabeçalho
    colunas_pend = [ws.cell(row=linha_cabecalho, column=c).value for c in range(1, 8)]
    indice_status = colunas_pend.index("Status") + 1

    pendentes = _montar_linhas_pendentes(resultado)
    cores_encontradas = set()
    for offset, item in enumerate(pendentes):
        celula = ws.cell(row=linha_cabecalho + 1 + offset, column=indice_status)
        assert celula.value == item["Status"]  # nunca substitui o texto original
        cor = celula.fill.fgColor.rgb[-6:] if celula.fill.fgColor.rgb else None
        cores_encontradas.add((item["Status"], cor))

    assert (STATUS_REVISAO_MANUAL, COR_LARANJA_FUNDO) in cores_encontradas
    assert (STATUS_SOMENTE_BANCO, COR_VERMELHO_FUNDO) in cores_encontradas


def test_secao_pendencias_com_zero_pendencias_mostra_faixa_verde(logger_silencioso, tmp_path):
    df_erp = construir_df_erp([
        {"data_usada": date(2026, 5, 10), "valor": 353.00, "favorecido": "IB Extintores - NF 2037"},
    ])
    df_banco = construir_df_banco([
        {"data": date(2026, 5, 10), "valor": -353.00, "favorecido": "PIX ENVIADO DES: IB EXTINTORES"},
    ])
    resultado = conciliar(df_erp, df_banco, logger_silencioso)
    assert (resultado["Status"] == STATUS_CONCILIADO).all()

    caminho = tmp_path / "Resultado.xlsx"
    exportar_resultado(resultado, caminho, logger_silencioso)
    workbook = openpyxl.load_workbook(caminho)
    ws = workbook[NOME_ABA_RESUMO]

    assert localizar_linha_com_texto(ws, MSG_SEM_PENDENCIAS) is not None


# --------------------------------------------------------------------------
# Base detalhada completa: filtros, congelamento e preservação dos dados.
# --------------------------------------------------------------------------

def test_base_detalhada_tem_filtro_e_congelamento_corretos(planilha_mista):
    workbook, resultado, _ = planilha_mista
    ws = workbook[NOME_ABA_BASE_DETALHADA]

    assert ws.freeze_panes == "A3"  # abaixo do título (linha 1) e cabeçalho (linha 2)
    total_colunas_visiveis = len(_colunas_visiveis(resultado))
    letra_final_esperada = ws.cell(row=2, column=total_colunas_visiveis).column_letter
    assert ws.auto_filter.ref == f"A2:{letra_final_esperada}{2 + len(resultado)}"


def test_base_detalhada_preserva_todas_as_colunas_visiveis_e_todas_as_linhas(planilha_mista):
    """Nenhuma linha é perdida, e nenhum valor das colunas visíveis é
    alterado — só as colunas de COLUNAS_OCULTAS_BASE_DETALHADA (pedido
    explícito do usuário, 2026-07-24) ficam de fora da planilha."""
    workbook, resultado, _ = planilha_mista
    ws = workbook[NOME_ABA_BASE_DETALHADA]

    linha_titulo = localizar_linha_com_texto(ws, TITULO_BASE_DETALHADA)
    assert linha_titulo == 1
    linha_cabecalho = 2

    colunas_esperadas = _colunas_visiveis(resultado)
    colunas_lidas = [
        ws.cell(row=linha_cabecalho, column=indice).value for indice in range(1, len(colunas_esperadas) + 1)
    ]
    assert colunas_lidas == colunas_esperadas

    for nome_oculto in COLUNAS_OCULTAS_BASE_DETALHADA:
        assert nome_oculto not in colunas_lidas

    for indice_linha in range(len(resultado)):
        linha_planilha = linha_cabecalho + 1 + indice_linha
        for indice_coluna, nome_coluna in enumerate(colunas_esperadas, start=1):
            valor_original = _normalizar_para_comparacao(resultado.iloc[indice_linha][nome_coluna])
            valor_lido = _normalizar_para_comparacao(ws.cell(row=linha_planilha, column=indice_coluna).value)
            assert valor_lido == valor_original, (
                f"Linha {indice_linha}, coluna '{nome_coluna}': esperado {valor_original!r}, lido {valor_lido!r}"
            )


def test_base_detalhada_nao_altera_status_original(planilha_mista):
    workbook, resultado, _ = planilha_mista
    ws = workbook[NOME_ABA_BASE_DETALHADA]
    colunas = _colunas_visiveis(resultado)
    indice_status = colunas.index("Status") + 1
    status_lidos = [ws.cell(row=3 + i, column=indice_status).value for i in range(len(resultado))]
    assert status_lidos == resultado["Status"].tolist()


def test_nenhuma_celula_mostra_nan_none_ou_nat(planilha_mista):
    workbook, _, _ = planilha_mista
    proibidos = {"nan", "none", "nat"}
    for nome_aba in workbook.sheetnames:
        ws = workbook[nome_aba]
        for linha in ws.iter_rows():
            for celula in linha:
                if isinstance(celula.value, str) and celula.value.strip().lower() in proibidos:
                    pytest.fail(f"[{nome_aba}] Célula {celula.coordinate} mostra valor técnico: {celula.value!r}")


def _normalizar_para_comparacao(valor):
    if valor is None:
        return None
    try:
        if pd.isna(valor):
            return None
    except (TypeError, ValueError):
        pass
    if isinstance(valor, datetime):
        return valor.date()
    if isinstance(valor, date):
        return valor
    if isinstance(valor, str):
        texto = valor.strip()
        return texto if texto else None
    if isinstance(valor, (int, float)):
        return round(float(valor), 2)
    return valor


# --------------------------------------------------------------------------
# Colunas opcionais (camada de IA) — não podem quebrar a exportação.
# --------------------------------------------------------------------------

def test_exportacao_funciona_com_colunas_extras_de_ia(logger_silencioso, tmp_path):
    """As 5 colunas de IA não quebram a exportação — e, desde 2026-07-24,
    ficam de fora da "Base Detalhada" (estão em COLUNAS_OCULTAS_BASE_DETALHADA)."""
    df_erp, df_banco = _construir_cenario_misto()
    resultado = conciliar(df_erp, df_banco, logger_silencioso)
    resultado = resultado.copy()
    resultado["Decisão IA"] = None
    resultado["Confiança IA"] = 0.0
    resultado["Motivo IA"] = None
    resultado["Validação IA"] = None
    resultado["Modelo IA"] = None

    caminho = tmp_path / "Resultado.xlsx"
    exportar_resultado(resultado, caminho, logger_silencioso)

    workbook = openpyxl.load_workbook(caminho)
    ws = workbook[NOME_ABA_BASE_DETALHADA]
    colunas_esperadas = _colunas_visiveis(resultado)
    colunas_lidas = [ws.cell(row=2, column=c).value for c in range(1, len(colunas_esperadas) + 1)]
    assert colunas_lidas == colunas_esperadas
    for nome_coluna_ia in ("Decisão IA", "Confiança IA", "Motivo IA", "Validação IA", "Modelo IA"):
        assert nome_coluna_ia not in colunas_lidas
    assert workbook.sheetnames == ["Resumo", "Base Detalhada"]
